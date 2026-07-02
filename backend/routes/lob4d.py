"""
LOB 4D Engine — fuente de verdad del cronograma/metrados para la simulación 4D.

Arquitectura (profesional y escalable):
  - El parseo pesado de los Excel corre AQUÍ (fuera del navegador), una sola vez,
    vía import explícito. Los datos persisten en Postgres AISLADOS por frente
    (model_urn), igual que inventory_assets.
  - La llave común es el CÓDIGO DE PARTIDA/EDT: los elementos del modelo llevan
    03_05_DSI_CodigoDePartida{1..4} en sus propiedades → el frontend une
    partida→elementos (misma lógica que budgetEngine) y colorea POR ELEMENTO.
  - Los Excel NO traen fechas explícitas: traen duración (EDT col CI) y avance
    real por periodo de valorización (VAL N°01..07). Las fechas se derivan de
    lob_config (fecha_inicio + dias_por_periodo) — ajustable, nunca inventado.

Fuentes parseadas:
  DURACIONES *.xlsm  hoja 'EDT'      → partida (X), desc (Y), unidad (Z), metrado (AA),
                                       P.U. (AB), rendimiento (AK), duración (CI),
                                       activity P6 (W), etiqueta de frente (M), orden (P)
  Metrados  *.xlsx   'CONTROL_OBRA'  → metrado contractual (E) / replanteado (F), P.U. (G)
                     'VALORIZACIONES'→ metrado ejecutado por periodo (E..K = VAL 1..7)
                     'MAPEO_FRENTES' → frente → prefijo de código base
"""
import io
import traceback
from flask import Blueprint, request, jsonify

lob4d_bp = Blueprint('lob4d', __name__)


# ─────────────────────────── Esquema ───────────────────────────

def ensure_lob4d_tables():
    try:
        from db import get_db_connection
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute("""
                CREATE TABLE IF NOT EXISTS lob_partidas (
                    model_urn    TEXT NOT NULL,
                    codigo       TEXT NOT NULL,
                    descripcion  TEXT,
                    unidad       TEXT,
                    metrado      DOUBLE PRECISION,
                    pu           DOUBLE PRECISION,
                    rendimiento  DOUBLE PRECISION,
                    duracion     DOUBLE PRECISION,
                    activity_id  TEXT,
                    frente_label TEXT,
                    orden        INTEGER,
                    updated_at   TIMESTAMP DEFAULT NOW(),
                    PRIMARY KEY (model_urn, codigo)
                )""")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS lob_avance (
                    model_urn   TEXT NOT NULL,
                    codigo      TEXT NOT NULL,
                    periodo     INTEGER NOT NULL,
                    metrado_ejec DOUBLE PRECISION,
                    PRIMARY KEY (model_urn, codigo, periodo)
                )""")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS lob_frentes (
                    model_urn TEXT NOT NULL,
                    frente    TEXT NOT NULL,
                    cod_base  TEXT NOT NULL,
                    PRIMARY KEY (model_urn, frente, cod_base)
                )""")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS lob_config (
                    model_urn        TEXT PRIMARY KEY,
                    fecha_inicio     DATE,
                    dias_por_periodo INTEGER DEFAULT 30,
                    updated_at       TIMESTAMP DEFAULT NOW()
                )""")
            conn.commit()
            print("[lob4d] Tablas LOB listas.")
    except Exception as e:
        print(f"[lob4d] ensure tables: {e}")


# ─────────────────────────── Parsers (puros, testeables sin BD) ───────────────────────────

def _num(v):
    try:
        if v is None or v == '':
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _txt(v):
    s = str(v).strip() if v is not None else ''
    return s or None


def parse_duraciones(file_bytes):
    """Hoja EDT → lista de partidas con duración/rendimiento. Solo filas con ITEM (col X)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if 'EDT' not in wb.sheetnames:
        wb.close()
        raise ValueError("El archivo de duraciones no tiene hoja 'EDT'.")
    ws = wb['EDT']
    out = {}
    # Cols (1-based): M=13 frente, P=16 orden, W=23 activity, X=24 item, Y=25 desc,
    #                 Z=26 unidad, AA=27 metrado, AB=28 pu, AK=37 rendimiento, CI=87 duración
    for row in ws.iter_rows(min_row=5, max_col=87, values_only=True):
        codigo = _txt(row[23])
        if not codigo:
            continue
        out[codigo] = {
            'codigo': codigo,
            'descripcion': _txt(row[24]),
            'unidad': _txt(row[25]),
            'metrado': _num(row[26]),
            'pu': _num(row[27]),
            'rendimiento': _num(row[36]),
            'duracion': _num(row[86]),
            'activity_id': _txt(row[22]),
            'frente_label': _txt(row[12]),
            'orden': int(_num(row[15]) or 0),
        }
    wb.close()
    return list(out.values())


def parse_metrados(file_bytes):
    """CONTROL_OBRA + VALORIZACIONES + MAPEO_FRENTES → metrados reales, avance y mapeo."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    control = {}
    if 'CONTROL_OBRA' in wb.sheetnames:
        for row in wb['CONTROL_OBRA'].iter_rows(min_row=2, max_col=9, values_only=True):
            codigo, tipo = _txt(row[0]), _txt(row[1])
            if not codigo or (tipo and tipo.upper() == 'TITULO'):
                continue
            control[codigo] = {
                'descripcion': _txt(row[2]),
                'unidad': _txt(row[3]),
                'metrado_contractual': _num(row[4]),
                'metrado': _num(row[5]) if _num(row[5]) is not None else _num(row[4]),
                'pu': _num(row[6]),
            }

    avance = {}
    if 'VALORIZACIONES' in wb.sheetnames:
        for row in wb['VALORIZACIONES'].iter_rows(min_row=5, max_col=12, values_only=True):
            codigo = _txt(row[0])
            # partidas reales tienen unidad (C); los títulos no
            if not codigo or not _txt(row[2]):
                continue
            periodos = {}
            for i in range(7):                      # E..K = periodos 1..7
                val = _num(row[4 + i])
                if val is not None and val != 0:
                    periodos[i + 1] = val
            if periodos:
                avance[codigo] = periodos

    frentes = []
    if 'MAPEO_FRENTES' in wb.sheetnames:
        for row in wb['MAPEO_FRENTES'].iter_rows(min_row=2, max_col=2, values_only=True):
            frente, cod = _txt(row[0]), _txt(row[1])
            if frente and cod:
                frentes.append({'frente': frente, 'cod_base': cod})

    wb.close()
    return {'control': control, 'avance': avance, 'frentes': frentes}


# ─────────────────────────── Endpoints ───────────────────────────

@lob4d_bp.route('/api/lob/import', methods=['POST'])
def lob_import():
    """Import explícito (re-import = refresh completo del frente). Multipart:
    model_urn, duraciones (.xlsm), metrados (.xlsx). Cualquiera de los dos archivos
    es opcional, pero al menos uno es requerido."""
    try:
        model_urn = request.form.get('model_urn') or request.args.get('model_urn')
        if not model_urn:
            return jsonify({'error': 'Falta model_urn (frente).'}), 400

        f_dur = request.files.get('duraciones')
        f_met = request.files.get('metrados')
        if not f_dur and not f_met:
            return jsonify({'error': 'Adjunta al menos un archivo (duraciones o metrados).'}), 400

        partidas = parse_duraciones(f_dur.read()) if f_dur else []
        met = parse_metrados(f_met.read()) if f_met else {'control': {}, 'avance': {}, 'frentes': []}

        # merge: CONTROL_OBRA manda en metrado/pu (replanteado real)
        by_code = {p['codigo']: p for p in partidas}
        for codigo, c in met['control'].items():
            p = by_code.setdefault(codigo, {
                'codigo': codigo, 'descripcion': None, 'unidad': None, 'metrado': None,
                'pu': None, 'rendimiento': None, 'duracion': None, 'activity_id': None,
                'frente_label': None, 'orden': 0,
            })
            p['descripcion'] = p['descripcion'] or c['descripcion']
            p['unidad'] = p['unidad'] or c['unidad']
            if c['metrado'] is not None:
                p['metrado'] = c['metrado']
            if c['pu'] is not None:
                p['pu'] = c['pu']

        from db import get_db_connection
        from psycopg2.extras import execute_values
        with get_db_connection() as conn:
            cur = conn.cursor()
            # refresh completo del frente (import explícito = fuente de verdad nueva)
            cur.execute("DELETE FROM lob_partidas WHERE model_urn = %s", (model_urn,))
            cur.execute("DELETE FROM lob_avance WHERE model_urn = %s", (model_urn,))
            cur.execute("DELETE FROM lob_frentes WHERE model_urn = %s", (model_urn,))

            if by_code:
                execute_values(cur, """
                    INSERT INTO lob_partidas (model_urn, codigo, descripcion, unidad, metrado,
                        pu, rendimiento, duracion, activity_id, frente_label, orden, updated_at)
                    VALUES %s""",
                    [(model_urn, p['codigo'], p['descripcion'], p['unidad'], p['metrado'],
                      p['pu'], p['rendimiento'], p['duracion'], p['activity_id'],
                      p['frente_label'], p['orden'], None) for p in by_code.values()],
                    template="(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())")

            rows_avance = [(model_urn, codigo, per, val)
                           for codigo, periodos in met['avance'].items()
                           for per, val in periodos.items()]
            if rows_avance:
                execute_values(cur, """
                    INSERT INTO lob_avance (model_urn, codigo, periodo, metrado_ejec)
                    VALUES %s""", rows_avance)

            if met['frentes']:
                execute_values(cur, """
                    INSERT INTO lob_frentes (model_urn, frente, cod_base) VALUES %s""",
                    [(model_urn, f['frente'], f['cod_base']) for f in met['frentes']])

            cur.execute("""
                INSERT INTO lob_config (model_urn) VALUES (%s)
                ON CONFLICT (model_urn) DO NOTHING""", (model_urn,))
            conn.commit()

        return jsonify({
            'status': 'ok',
            'partidas': len(by_code),
            'partidas_con_avance': len(met['avance']),
            'frentes_mapeados': len(met['frentes']),
        }), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@lob4d_bp.route('/api/lob/timeline', methods=['GET'])
def lob_timeline():
    """Todo lo que el motor 4D del frontend necesita, en UNA llamada:
    config + partidas (metrado/duración/pu) + avance por periodo + mapeo de frentes.
    El cruce partida→elementos lo hace el cliente (ya tiene el inventario en memoria)."""
    try:
        model_urn = request.args.get('model_urn')
        if not model_urn:
            return jsonify({'error': 'Falta model_urn.'}), 400
        prefix = request.args.get('prefix')  # opcional: filtrar por cod_base (frente físico)

        from db import get_db_connection
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute("""
                SELECT fecha_inicio::text, dias_por_periodo FROM lob_config
                WHERE model_urn = %s""", (model_urn,))
            row = cur.fetchone()
            config = {'fecha_inicio': row[0] if row else None,
                      'dias_por_periodo': (row[1] if row else None) or 30}

            q = """SELECT codigo, descripcion, unidad, metrado, pu, rendimiento,
                          duracion, activity_id, frente_label, orden
                   FROM lob_partidas WHERE model_urn = %s"""
            params = [model_urn]
            if prefix:
                q += " AND codigo LIKE %s"
                params.append(prefix + '%')
            q += " ORDER BY orden, codigo"
            cur.execute(q, params)
            partidas = [{
                'codigo': r[0], 'descripcion': r[1], 'unidad': r[2], 'metrado': r[3],
                'pu': r[4], 'rendimiento': r[5], 'duracion': r[6], 'activity_id': r[7],
                'frente_label': r[8], 'orden': r[9],
            } for r in cur.fetchall()]

            cur.execute("""
                SELECT codigo, periodo, metrado_ejec FROM lob_avance
                WHERE model_urn = %s ORDER BY codigo, periodo""", (model_urn,))
            avance = {}
            for codigo, per, val in cur.fetchall():
                avance.setdefault(codigo, {})[str(per)] = val

            cur.execute("""
                SELECT frente, cod_base FROM lob_frentes WHERE model_urn = %s
                ORDER BY frente, cod_base""", (model_urn,))
            frentes = {}
            for frente, cod in cur.fetchall():
                frentes.setdefault(frente, []).append(cod)

        return jsonify({'config': config, 'partidas': partidas,
                        'avance': avance, 'frentes': frentes}), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@lob4d_bp.route('/api/lob/config', methods=['POST'])
def lob_set_config():
    """Fija la fecha de inicio del proyecto y días por periodo de valorización
    (lo único que los Excel no traen — explícito, no inventado)."""
    try:
        data = request.get_json() or {}
        model_urn = data.get('model_urn')
        if not model_urn:
            return jsonify({'error': 'Falta model_urn.'}), 400
        from db import get_db_connection
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO lob_config (model_urn, fecha_inicio, dias_por_periodo, updated_at)
                VALUES (%s, %s, %s, NOW())
                ON CONFLICT (model_urn) DO UPDATE SET
                    fecha_inicio = COALESCE(EXCLUDED.fecha_inicio, lob_config.fecha_inicio),
                    dias_por_periodo = COALESCE(EXCLUDED.dias_por_periodo, lob_config.dias_por_periodo),
                    updated_at = NOW()""",
                (model_urn, data.get('fecha_inicio'), data.get('dias_por_periodo')))
            conn.commit()
        return jsonify({'status': 'ok'}), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
