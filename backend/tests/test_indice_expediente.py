"""El indice del expediente: la relacion de lo entregado, y la salida de la jaula.

POR QUE ESTOS TESTS
-------------------
Dos cosas tienen que seguir siendo ciertas pase lo que pase:

1. El indice NO incluye el trabajo en curso salvo que se pida a proposito. Un WIP
   no se ha comprometido con nadie; colarlo en la relacion de entregados da una
   foto falsa de lo entregado, que es exactamente lo que una supervision castiga.
2. La hoja se abre sin esta plataforma. Es la respuesta a "¿y si me voy de tu
   ECD?", asi que si un dia deja de poder reabrirse, el argumento se cae.
"""
import io

import indice_expediente as ie


def test_no_incluye_el_trabajo_en_curso_por_defecto():
    assert 'WIP' not in ie.ESTADOS_ENTREGADOS
    assert set(ie.ESTADOS_ENTREGADOS) == {'SHARED', 'PUBLISHED', 'ARCHIVED'}


def test_el_resumen_cuenta_lo_que_falta():
    """Lo que importa antes de una entrega es lo que NO consta."""
    filas = [
        {'estado': 'PUBLISHED', 'idoneidad': 'A1', 'revision': 'C01', 'emitida_por': 'Ana'},
        {'estado': 'PUBLISHED', 'idoneidad': None, 'revision': None, 'emitida_por': None},
        {'estado': 'SHARED', 'idoneidad': 'S3', 'revision': 'C02', 'emitida_por': None},
    ]
    r = ie.resumen(filas)
    assert r['documentos'] == 3
    assert r['sin_idoneidad'] == 1
    assert r['sin_revision'] == 1
    assert r['sin_emisor'] == 2
    assert r['por_estado'] == {'PUBLISHED': 2, 'SHARED': 1}


def test_la_hoja_se_reabre_y_lleva_de_que_obra_es():
    """Sin obra ni fecha, una hoja suelta no vale como evidencia."""
    import openpyxl
    filas = [{c: None for c, _e in ie.COLUMNAS} | {'codigo': '500125-X', 'estado': 'PUBLISHED'}]
    datos = ie.a_excel(filas, 'proyectos/OBRA_X', generado_por='Ana', estados=None)
    ws = openpyxl.load_workbook(io.BytesIO(datos)).active

    texto = '\n'.join(str(c.value) for fila in ws.iter_rows() for c in fila if c.value)
    assert 'proyectos/OBRA_X' in texto
    assert 'Ana' in texto
    assert 'trabajo en curso' in texto      # el alcance se declara en la hoja
    assert '500125-X' in texto


def test_la_hoja_no_lleva_formulas_ni_macros():
    """Una entidad publica no acepta un fichero que ejecuta cosas."""
    import openpyxl
    datos = ie.a_excel([], 'proyectos/OBRA_X', generado_por=None, estados=None)
    wb = openpyxl.load_workbook(io.BytesIO(datos))
    for fila in wb.active.iter_rows():
        for c in fila:
            assert not (isinstance(c.value, str) and c.value.startswith('=')), c.value
    assert not getattr(wb, 'vba_archive', None)


# ── La columna «Ubicacion en el ECD» ──────────────────────────────────────
#
# El indice del expediente es el documento que se entrega para que OTRO
# encuentre los ficheros. Se quitaba siempre el primer tramo de la ruta, dando
# por hecho que el arbol cuelga de una raiz interna. En un arbol sin ella eso se
# comia la carpeta de primer nivel: un documento de «01. PLANOS/DRENAJE» salia
# listado en «DRENAJE». Mandar a alguien a una carpeta que no existe es peor que
# no decirle nada.

class CursorDeArbol:
    """Devuelve las filas del indice y cuantas raices tiene la obra."""

    def __init__(self, filas, raices):
        self.filas = filas
        self.raices = raices
        self._r = []
        self._modo = None

    def execute(self, sql, params=None):
        if 'count(*)' in sql.lower():
            self._modo = 'raices'
        else:
            self._modo = 'filas'

    def fetchall(self):
        # La ultima columna es `fn.id`: el indice la lleva desde que se filtra
        # por el permiso documental de cada documento.
        return [(nombre, ruta, 'PUBLISHED', 'A1', 'C01', 1,
                 None, None, None, None, True, 1048576,
                 'nodo-%d' % i)
                for i, (nombre, ruta) in enumerate(self.filas)]

    def fetchone(self):
        return (self.raices,)


def _ruta_de(filas, raices):
    cur = CursorDeArbol(filas, raices)
    return [f['ruta'] for f in ie.filas_del_indice(cur, 'obra/X')]


def test_con_raiz_unica_se_quita_la_raiz_pero_no_la_carpeta():
    """El nombre del contenedor raiz no le dice nada a quien lee el indice."""
    filas = [('PL-001.pdf', 'PQT8_TALARA/01. PLANOS/DRENAJE/PL-001.pdf')]
    assert _ruta_de(filas, raices=1) == ['01. PLANOS/DRENAJE']


def test_sin_raiz_unica_la_carpeta_de_primer_nivel_se_conserva():
    """Este era el fallo: '01. PLANOS' desaparecia de la ubicacion."""
    filas = [('PL-001.pdf', '01. PLANOS/DRENAJE/PL-001.pdf')]
    assert _ruta_de(filas, raices=3) == ['01. PLANOS/DRENAJE']


def test_un_documento_en_la_propia_raiz_no_inventa_carpeta():
    filas = [('PL-001.pdf', 'PQT8_TALARA/PL-001.pdf')]
    assert _ruta_de(filas, raices=1) == ['']


def test_el_nombre_del_fichero_nunca_va_en_la_ubicacion():
    """Va en su columna. Repetirlo en la ubicacion hace creer que hay una
    carpeta con ese nombre."""
    filas = [('PL-001.pdf', '01. PLANOS/PL-001.pdf')]
    assert 'PL-001.pdf' not in _ruta_de(filas, raices=2)[0]
