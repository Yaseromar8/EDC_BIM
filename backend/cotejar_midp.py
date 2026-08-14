# -*- coding: utf-8 -*-
"""Coteja el MIDP de una obra contra lo que hay de verdad en el ECD.

POR QUÉ EXISTE
--------------
El MIDP dice qué se comprometió a entregar. El ECD dice qué hay. Hasta ahora
nadie había puesto las dos listas una al lado de la otra, y al hacerlo por
primera vez en PQT8 salió esto: de 282 planos aprobados, 0 estaban en el ECD.
El expediente vivía en una carpeta de Drive y la plataforma tenía las fotos.

NO SUBE NADA NI ESCRIBE NADA. Solo lee y compara. Lo que produce es la foto que
hace falta para decidir: qué falta, qué sobra, y con qué convención de nombres
trabaja esta obra de verdad.

MULTI-OBRA A PROPÓSITO
----------------------
Ni la ruta, ni el patrón, ni los códigos están cableados: todo sale del MIDP que
se le pase. Toda obra ISO 19650 tiene uno, así que "importa tu MIDP" sirve de
alta para cualquier obra nueva, no solo para Talara. Si algún día esto se cablea
a los valores de PQT8, deja de ser producto y pasa a ser un apaño.

USO
    python cotejar_midp.py --midp "<ruta al xlsx>" [--carpeta "<ruta a planos>"]
                           [--obra <model_urn>]
"""
import argparse
import collections
import os
import re
import sys

# La consola de Windows usa cp1252 y revienta con cualquier acento o flecha.
# Sin esto el informe no se puede ni imprimir en la maquina donde se usa.
try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:                                      # pragma: no cover
    pass

try:
    import openpyxl
except ImportError:                                    # pragma: no cover
    openpyxl = None


# Las columnas del MIDP de la Guía Nacional BIM (anexo 6.3). Se detectan por
# CABECERA, no por posición fija: dos obras pueden tener columnas de más o de
# menos, y un índice cableado se rompe con el primer MIDP ajeno.
CABECERAS = {
    'titulo': ('titulo del contenedor', 'título del contenedor'),
    'formato': ('formato o extensión', 'formato o extension'),
    'proyecto': ('código de proyecto', 'codigo de proyecto'),
    'originador': ('código de originador', 'codiggo originador', 'código originador'),
    'volumen': ('código de volumen', 'codigo de volumen'),
    'nivel': ('código de nivel', 'codigo de nivel'),
    'tipo': ('código de tipo', 'codigo de tipo'),
    'disciplina': ('código de disciplina', 'codigo de disciplina'),
    'numeracion': ('código de numeración', 'codigo de numeracion'),
    'estado': ('código de estado', 'codigo de estado'),
    'revision': ('revisión', 'revision'),
    'identificador': ('identificación del contenedor', 'identificacion del contenedor'),
    # Lo que convierte el plan en un plan: QUIEN y PARA CUANDO. Sin fecha no hay
    # "vencido", y sin vencido esto es un inventario, no un plan de entrega.
    # El TIDP los trae siempre; el MIDP, segun como lo haya montado la obra.
    'responsable': ('responsable', 'equipo responsable', 'tarea', 'equipo de tarea',
                    'task team', 'autor'),
    'fecha': ('fecha de entrega', 'fecha comprometida', 'fecha', 'entrega',
              'fecha de emisión', 'fecha de emision'),
    'hito': ('hito', 'milestone', 'etapa', 'fase de entrega'),
}


def _norm(v):
    return ' '.join(str(v or '').split()).strip().lower()


def localizar_columnas(ws, hasta_fila=8):
    """Mapa campo -> índice de columna, buscando por el texto de la cabecera."""
    cols = {}
    for r in range(1, hasta_fila + 1):
        for c in range(1, ws.max_column + 1):
            t = _norm(ws.cell(row=r, column=c).value)
            if not t:
                continue
            for campo, claves in CABECERAS.items():
                if campo in cols:
                    continue
                if any(t.startswith(k) for k in claves):
                    cols[campo] = c
    return cols


def leer_midp(ruta):
    wb = openpyxl.load_workbook(ruta, data_only=True)
    entregables = []
    for hoja in wb.sheetnames:
        ws = wb[hoja]
        cols = localizar_columnas(ws)
        if 'identificador' not in cols:
            continue
        for r in range(2, ws.max_row + 1):
            fila = {k: ws.cell(row=r, column=c).value for k, c in cols.items()}
            ident = str(fila.get('identificador') or '').strip()
            if not ident or ident.lower().startswith('identific'):
                continue
            fila['hoja'] = hoja
            fila['identificador'] = ident
            entregables.append(fila)
    wb.close()
    return entregables


def patron_desde(entregables):
    """Deduce la convención de nombres a partir de los identificadores reales.

    Se construye con los VALORES que la obra usa de verdad (TP|HD|ST...), no con
    un comodín. Un patrón laxo acepta cualquier cosa y deja de ser un control.
    """
    campos = ['proyecto', 'originador', 'volumen', 'nivel', 'tipo', 'disciplina', 'numeracion']
    vistos = {}
    for k in campos:
        vals = {str(e.get(k) or '').strip().upper() for e in entregables}
        vistos[k] = sorted(v for v in vals if v)
    partes = []
    for k in campos:
        vs = vistos[k]
        if not vs:
            partes.append(r'[A-Z0-9]+')
        elif len(vs) <= 12 and all(re.fullmatch(r'[A-Z0-9]+', v) for v in vs):
            partes.append('(?:%s)' % '|'.join(vs))
        else:
            largos = {len(v) for v in vs}
            partes.append(r'\d{%d}' % largos.pop() if len(largos) == 1 and vs[0].isdigit()
                          else r'[A-Z0-9]+')
    return '^' + '-'.join(partes) + '$', vistos


def frecuencias(entregables, campo, tope=12):
    c = collections.Counter(str(e.get(campo) or '').strip() for e in entregables)
    c.pop('', None)
    return c.most_common(tope)


def nombres_del_ecd(obra):
    # Autosuficiente: carga el entorno y levanta el pool por su cuenta, como el
    # resto de guiones del backend. Si depende de que alguien lo inicialice
    # antes, no se puede lanzar suelto — que es justo para lo que sirve.
    aqui = os.path.dirname(os.path.abspath(__file__))
    sys.path.insert(0, aqui)
    try:
        from dotenv import load_dotenv
        load_dotenv(os.path.join(os.path.dirname(aqui), '.env'))
    except ImportError:                                # pragma: no cover
        pass
    import db as _db
    if getattr(_db, 'db_pool', None) is None:
        _db.init_db_pool()
    from db import get_db_connection
    with get_db_connection() as conn:
        cur = conn.cursor()
        if obra:
            cur.execute("""SELECT name, status FROM file_nodes
                            WHERE node_type='FILE' AND COALESCE(is_deleted,false)=false
                              AND model_urn = %s""", (obra,))
        else:
            cur.execute("""SELECT name, status FROM file_nodes
                            WHERE node_type='FILE' AND COALESCE(is_deleted,false)=false""")
        return [(r[0], r[1]) for r in cur.fetchall()]


def base(nombre):
    return os.path.splitext(str(nombre or '').strip())[0].upper()


def main():
    ap = argparse.ArgumentParser(description='Coteja el MIDP contra el ECD. No escribe nada.')
    ap.add_argument('--midp', required=True, help='ruta del MIDP (.xlsx)')
    ap.add_argument('--carpeta', help='carpeta de entregables en disco (opcional)')
    ap.add_argument('--obra', help='model_urn de la obra (por defecto, todas)')
    a = ap.parse_args()

    if openpyxl is None:
        print('Falta openpyxl:  pip install openpyxl'); return 1

    print('=' * 74)
    print('COTEJO MIDP  ↔  ECD          (solo lectura: no sube ni modifica nada)')
    print('=' * 74)

    ent = leer_midp(a.midp)
    print('\nMIDP: %d entregables comprometidos' % len(ent))
    for campo in ('disciplina', 'tipo', 'estado', 'revision'):
        f = frecuencias(ent, campo)
        if f:
            print('  %-11s %s' % (campo, ', '.join('%s×%d' % (v, n) for v, n in f)))

    patron, vistos = patron_desde(ent)
    print('\nCONVENCIÓN deducida de los nombres reales:')
    print('  %s' % patron)
    print('  (esto es lo que habría que guardar como patrón DE ESTA OBRA)')

    ids = {base(e['identificador']) for e in ent}

    if a.carpeta:
        disco = {}
        for r, _d, fs in os.walk(a.carpeta):
            for f in fs:
                if f.lower().endswith('.pdf'):
                    disco[base(f)] = os.path.join(r, f)
        print('\nCARPETA: %d PDF' % len(disco))
        print('  coinciden con el MIDP : %d' % len(ids & set(disco)))
        print('  en disco y NO en MIDP : %d' % len(set(disco) - ids))

    ecd = nombres_del_ecd(a.obra)
    en_ecd = {base(n) for n, _s in ecd}
    dentro = ids & en_ecd
    print('\nECD: %d ficheros' % len(ecd))
    print('  comprometidos que SÍ están : %d de %d' % (len(dentro), len(ids)))
    print('  comprometidos que FALTAN   : %d' % len(ids - en_ecd))

    if dentro:
        est = collections.Counter(s or 'sin estado' for n, s in ecd if base(n) in dentro)
        print('  estado de los que están    : %s'
              % ', '.join('%s×%d' % (k, v) for k, v in est.most_common()))

    faltan = sorted(ids - en_ecd)
    if faltan:
        print('\n  Primeros que faltan:')
        for i in faltan[:8]:
            print('    %s' % i)
        if len(faltan) > 8:
            print('    … y %d más' % (len(faltan) - 8))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
