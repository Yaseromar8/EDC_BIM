"""
Script de Ingesta del Presupuesto Maestro
==========================================
Lee Presupuesto_Rv01.xlsx y lo guarda en la tabla `presupuesto_maestro` de PostgreSQL.
Puede ejecutarse directamente o ser invocado desde un endpoint API.
"""
import openpyxl
import json
import sys
import os

# -------------------------------------------------------------------
# 1. LECTURA DEL EXCEL
# -------------------------------------------------------------------
def parse_presupuesto_excel(filepath, sheet_name='PPTO'):
    """
    Lee el Excel y devuelve una lista de diccionarios con la estructura:
    {item, descripcion, unidad, metrado_contractual, precio_unitario, parcial_contractual, nivel, es_titulo, parent_item}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        item_raw = row[0]
        if item_raw is None:
            continue
        item = str(item_raw).strip()
        if not item or item == 'None':
            continue
        
        # Columnas: 0=ITEM, 1=CUENTA COSTOS (skip), 2=DESCRIPCION, 3=UND, 4=METRADO, 5=P.U, 6=P.PARCIAL
        descripcion = str(row[2]).strip() if row[2] else ''
        unidad = str(row[3]).strip() if row[3] else ''
        metrado_contractual = float(row[4]) if row[4] is not None else 0.0
        precio_unitario = float(row[5]) if row[5] is not None else 0.0
        parcial_contractual = float(row[6]) if row[6] is not None else 0.0
        
        # Calcular nivel jerárquico (06.01.03 = nivel 3)
        parts = item.split('.')
        nivel = len(parts)
        
        # Determinar si es título (carpeta) o partida ejecutable
        es_titulo = (unidad == '')
        
        # Calcular parent_item (06.01.03.06 -> parent = 06.01.03)
        parent_item = '.'.join(parts[:-1]) if len(parts) > 1 else None
        
        rows.append({
            'item': item,
            'descripcion': descripcion,
            'unidad': unidad if not es_titulo else None,
            'metrado_contractual': metrado_contractual if not es_titulo else None,
            'precio_unitario': precio_unitario if not es_titulo else None,
            'parcial_contractual': parcial_contractual,
            'nivel': nivel,
            'es_titulo': es_titulo,
            'parent_item': parent_item
        })
    
    wb.close()
    return rows


# -------------------------------------------------------------------
# 2. CREACIÓN DE TABLA EN POSTGRESQL
# -------------------------------------------------------------------
CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS presupuesto_maestro (
    id SERIAL PRIMARY KEY,
    model_urn TEXT NOT NULL,
    item TEXT NOT NULL,
    descripcion TEXT,
    unidad TEXT,
    metrado_contractual NUMERIC,
    precio_unitario NUMERIC,
    parcial_contractual NUMERIC,
    nivel INTEGER NOT NULL DEFAULT 1,
    es_titulo BOOLEAN NOT NULL DEFAULT FALSE,
    parent_item TEXT,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(model_urn, item)
);

CREATE INDEX IF NOT EXISTS idx_presupuesto_model_urn ON presupuesto_maestro(model_urn);
CREATE INDEX IF NOT EXISTS idx_presupuesto_parent ON presupuesto_maestro(parent_item);
CREATE INDEX IF NOT EXISTS idx_presupuesto_item ON presupuesto_maestro(item);
"""


def create_table(conn):
    """Crea la tabla presupuesto_maestro si no existe."""
    cursor = conn.cursor()
    cursor.execute(CREATE_TABLE_SQL)
    conn.commit()
    print("[Presupuesto] Tabla presupuesto_maestro creada/verificada OK")


# -------------------------------------------------------------------
# 3. INSERCIÓN MASIVA
# -------------------------------------------------------------------
def insert_presupuesto(conn, model_urn, rows):
    """Inserta (o reemplaza) las filas del presupuesto para un model_urn."""
    cursor = conn.cursor()
    
    # Limpiar datos anteriores de este proyecto
    cursor.execute("DELETE FROM presupuesto_maestro WHERE model_urn = %s", (model_urn,))
    deleted = cursor.rowcount
    if deleted > 0:
        print(f"[Presupuesto] Limpiados {deleted} registros anteriores para {model_urn}")
    
    # Inserción masiva
    from psycopg2.extras import execute_values
    insert_sql = """
        INSERT INTO presupuesto_maestro 
            (model_urn, item, descripcion, unidad, metrado_contractual, precio_unitario, parcial_contractual, nivel, es_titulo, parent_item)
        VALUES %s
    """
    records = [
        (model_urn, r['item'], r['descripcion'], r['unidad'],
         r['metrado_contractual'], r['precio_unitario'], r['parcial_contractual'],
         r['nivel'], r['es_titulo'], r['parent_item'])
        for r in rows
    ]
    
    execute_values(cursor, insert_sql, records)
    conn.commit()
    
    titulos = sum(1 for r in rows if r['es_titulo'])
    partidas = len(rows) - titulos
    print(f"[Presupuesto] Insertados {len(rows)} registros ({titulos} títulos + {partidas} partidas)")
    return len(rows)


# -------------------------------------------------------------------
# 4. EJECUCIÓN DIRECTA (CLI)
# -------------------------------------------------------------------
if __name__ == '__main__':
    excel_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'Presupuesto_Rv01.xlsx')
    
    if not os.path.exists(excel_path):
        print(f"ERROR: No se encuentra {excel_path}")
        sys.exit(1)
    
    print(f"[Presupuesto] Leyendo Excel: {excel_path}")
    rows = parse_presupuesto_excel(excel_path)
    print(f"[Presupuesto] Total filas parseadas: {len(rows)}")
    
    # Estadísticas
    niveles = {}
    for r in rows:
        niveles[r['nivel']] = niveles.get(r['nivel'], 0) + 1
    
    print("\n=== ESTADÍSTICAS ===")
    for n in sorted(niveles.keys()):
        print(f"  Nivel {n}: {niveles[n]} filas")
    
    titulos = sum(1 for r in rows if r['es_titulo'])
    print(f"  Títulos (carpetas): {titulos}")
    print(f"  Partidas (ejecutables): {len(rows) - titulos}")
    
    # Intentar conexión a BD
    try:
        sys.path.insert(0, os.path.dirname(__file__))
        from db import get_db_connection
        
        # Usar el model_urn del proyecto activo, o uno por defecto
        model_urn = sys.argv[1] if len(sys.argv) > 1 else 'global'
        
        with get_db_connection() as conn:
            create_table(conn)
            count = insert_presupuesto(conn, model_urn, rows)
            print(f"\n[Presupuesto] ¡COMPLETADO! {count} registros en presupuesto_maestro")
    
    except Exception as e:
        print(f"\n[Presupuesto] No se pudo conectar a PostgreSQL: {e}")
        print("[Presupuesto] Guardando como JSON local para verificación...")
        
        output_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'presupuesto_maestro.json')
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(rows, f, ensure_ascii=False, indent=2)
        print(f"[Presupuesto] JSON guardado en: {output_path}")
        print(f"[Presupuesto] Total: {len(rows)} filas")
