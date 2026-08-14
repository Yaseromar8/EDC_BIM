# -*- coding: utf-8 -*-
"""Lee el MIDP/TIDP en el formato de la Guía Nacional BIM del Perú.

POR QUÉ NO VALE EL LECTOR GENÉRICO
----------------------------------
`cotejar_midp.leer_midp` busca cada campo por su cabecera y se queda con la
primera coincidencia. Con esta plantilla eso da un plan equivocado, y en
silencio:

  · La cabecera es de CUATRO niveles (filas 4, 6, 7 y 8), no de uno.
  · Hay TRES bloques de columnas, uno por etapa RIBA (3A, 3B, 4), y cada uno
    repite «Parte Responsable», «LOIN», «Tiempo estimado» y «Fecha de Entrega».
    Coger la primera «Fecha de Entrega» es coger la de RIBA 3A, que en el MIDP
    de PQT8 está VACÍA: el plan entraría entero sin fechas, y sin fecha no hay
    entregable vencido.
  · «Parte Responsable» no empieza por «responsable», así que ni se detectaba.

LOS RANGOS CON @
----------------
Una fila puede comprometer MUCHOS documentos:

    500125-CSSP001-740-XX-DR-ST-004120@004145

son los planos 004120 a 004145 — veintiséis, no uno. En el TIDP de PQT8, 36 de
las 118 filas son rangos, así que leerlas como una sola comprometería 118
entregables cuando de verdad son 311. Un plan que declara la tercera parte de lo
que se prometió no sirve para controlar nada.

Se expanden aquí, al leer, porque el resto del sistema razona por documento: un
rango no se puede vincular a un PDF, ni marcar entregado a medias.
"""

import re
import warnings

# openpyxl avisa de que pierde formas y dibujos. Da igual: aquí se leen valores.
warnings.filterwarnings('ignore', module='openpyxl')

RANGO = re.compile(r'^(.*?)(\d+)@(\d+)$')

# Etiquetas que buscamos, normalizadas. Se comparan por «contiene», no por
# «empieza por»: la plantilla escribe «Parte Responsable» y «Código de estado».
CAMPOS_BASE = {
    'paquete': ('paquete',),
    'componente': ('componente',),
    'titulo': ('titulo del contenedor', 'título del contenedor'),
    'descripcion': ('descripción del contenedor', 'descripcion del contenedor'),
    'formato': ('formato o extensión', 'formato o extension'),
    'escala': ('escala',),
    'identificador': ('identificación del contenedor', 'identificacion del contenedor'),
}
PARTES_CODIGO = {
    'proyecto': ('código de proyecto', 'codigo de proyecto'),
    'originador': ('códiggo originador', 'código originador', 'codigo originador',
                   'código de originador'),
    'volumen': ('código de volumen', 'codigo de volumen'),
    'nivel': ('código de nivel', 'codigo de nivel'),
    'tipo': ('código de tipo', 'codigo de tipo'),
    'disciplina': ('código de disciplina', 'codigo de disciplina'),
    'numeracion': ('código de numeración', 'codigo de numeracion'),
    'estado': ('código de estado', 'codigo de estado'),
    'revision': ('revisión', 'revision'),
}
CAMPOS_ETAPA = {
    'responsable': ('parte responsable',),
    'lod': ('lod',),
    'loi': ('loi',),
    'produccion': ('tiempo estimado',),
    'fecha': ('fecha de entrega',),
}


def _txt(ws, r, c):
    v = ws.cell(row=r, column=c).value
    return ' '.join(str(v).split()) if v is not None else ''


def _norm(s):
    return (s or '').strip().lower()


def _localizar(ws, filas_cabecera=10):
    """Devuelve (col_base, col_etapa, primera_fila_de_datos).

    col_base:  campo -> columna
    col_etapa: etapa -> {campo -> columna}
    """
    col_base, col_etapa = {}, {}
    fila_ident = None

    # 1. La fila de campos es la que lleva «Identificación del contenedor».
    for r in range(1, filas_cabecera + 1):
        for c in range(1, ws.max_column + 1):
            t = _norm(_txt(ws, r, c))
            if any(k in t for k in CAMPOS_BASE['identificador']):
                fila_ident = r
                break
        if fila_ident:
            break
    if not fila_ident:
        return None, None, None

    # 2. Campos base y partes del código, en la fila de campos y la siguiente.
    for r in (fila_ident, fila_ident + 1):
        for c in range(1, ws.max_column + 1):
            t = _norm(_txt(ws, r, c))
            if not t:
                continue
            for campo, claves in {**CAMPOS_BASE, **PARTES_CODIGO}.items():
                if campo not in col_base and any(k in t for k in claves):
                    col_base[campo] = c

    # 3. Bloques por etapa. La etiqueta de etapa («Etapa RIBA 4») aparece UNA vez,
    #    en la primera columna del bloque, y vale hasta la siguiente: se arrastra.
    fila_etapa = max(1, fila_ident - 3)
    fila_campo_etapa = max(1, fila_ident - 1)
    etapa_actual = None
    for c in range(1, ws.max_column + 1):
        et = _norm(_txt(ws, fila_etapa, c))
        if 'etapa' in et:
            etapa_actual = _txt(ws, fila_etapa, c)
            col_etapa.setdefault(etapa_actual, {})
        if not etapa_actual:
            continue
        for r in (fila_campo_etapa, fila_ident):
            t = _norm(_txt(ws, r, c))
            if not t:
                continue
            for campo, claves in CAMPOS_ETAPA.items():
                if campo not in col_etapa[etapa_actual] and any(k in t for k in claves):
                    col_etapa[etapa_actual][campo] = c

    return col_base, col_etapa, fila_ident + 2


def expandir_rango(identificador):
    """'…-004120@004145' -> ['…-004120', …, '…-004145']. Sin @ devuelve [tal cual].

    Si el rango va al revés o es absurdamente grande NO se inventa nada: se
    devuelve el identificador original y que se vea en el plan tal como está en
    el Excel. Expandir 900.000 filas por una errata sería peor que no expandir.
    """
    ident = (identificador or '').strip()
    m = RANGO.match(ident)
    if not m:
        return [ident] if ident else []
    prefijo, desde, hasta = m.group(1), m.group(2), m.group(3)
    ancho = len(desde)
    try:
        a, b = int(desde), int(hasta)
    except ValueError:
        return [ident]
    if b < a or (b - a) > 2000:
        return [ident]
    return [f'{prefijo}{str(n).zfill(ancho)}' for n in range(a, b + 1)]


def leer(ruta):
    """Lista de entregables, con los rangos ya expandidos.

    Cada elemento es un compromiso: un documento, una etapa, una fecha.
    """
    import openpyxl
    wb = openpyxl.load_workbook(ruta, data_only=True)
    salida = []
    for ws in wb.worksheets:
        col_base, col_etapa, fila0 = _localizar(ws)
        if not col_base or 'identificador' not in col_base:
            continue
        for r in range(fila0, ws.max_row + 1):
            ident = _txt(ws, r, col_base['identificador'])
            if not ident or _norm(ident).startswith('identific'):
                continue
            base = {k: _txt(ws, r, c) for k, c in col_base.items()}

            # Una fila puede tener varias etapas rellenas: cada una es una
            # entrega distinta del mismo contenedor, con su propia fecha.
            etapas = []
            for etapa, cols in (col_etapa or {}).items():
                valores = {k: _txt(ws, r, c) for k, c in cols.items()}
                # Se considera que la etapa aplica si dice QUIEN o CUANDO.
                if valores.get('responsable') or valores.get('fecha'):
                    etapas.append((etapa, valores))
            if not etapas:
                etapas = [(None, {})]

            for etapa, val in etapas:
                for uno in expandir_rango(ident):
                    e = dict(base)
                    e['identificador'] = uno
                    e['hoja'] = ws.title
                    e['hito'] = etapa or base.get('paquete') or None
                    e['responsable'] = val.get('responsable') or None
                    e['fecha'] = _fecha_cruda(ws, r, (col_etapa.get(etapa) or {}).get('fecha'))
                    e['lod'] = val.get('lod') or None
                    e['loi'] = val.get('loi') or None
                    e['de_rango'] = uno != ident
                    salida.append(e)
    wb.close()
    return salida


def _fecha_cruda(ws, r, c):
    """El valor SIN convertir a texto: openpyxl ya devuelve datetime si lo es.

    Pasar por str() y volver a parsear pierde el tipo y abre la puerta a leer
    03/04 como 3 de abril o 4 de marzo segun el idioma del que mire.
    """
    if not c:
        return None
    return ws.cell(row=r, column=c).value
