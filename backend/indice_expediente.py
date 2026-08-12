# -*- coding: utf-8 -*-
"""El indice del expediente: todo lo publicado de una obra, en una sola tabla.

POR QUE EXISTE
--------------
Es lo primero que pide una supervision, y hasta ahora no habia forma de sacarlo.
El portal navega carpeta por carpeta (/api/docs/list) y el buscador corta en 50
resultados, asi que para responder "dame la relacion de todo lo aprobado, con su
revision, su codigo de idoneidad, la fecha y quien lo aprobo" habia que abrir
carpetas a mano y apuntar. Con 282 planos, eso no se hace.

QUE RESUELVE ADEMAS
-------------------
La entidad pregunta siempre lo mismo antes de firmar: "y si manana me voy de tu
plataforma, ¿como saco MI expediente?". Un ECD del que no se puede salir es una
jaula, y se nota. Este indice es la respuesta: una hoja de calculo que se abre en
cualquier ordenador dentro de diez anos, sin esta plataforma y sin internet.

NO ES UN INFORME BONITO. Es la relacion de contenedores de informacion con los
campos que la ISO 19650 pide identificar en cada uno. Si algo no consta, sale
vacio y se ve: un hueco visible es informacion, un hueco relleno con "N/A" es
ruido.
"""
import io
from datetime import datetime, timezone

# Las columnas del indice, en el orden en que las lee una supervision: primero
# QUE es y DONDE esta, luego en que estado esta y con que autoridad, y al final
# quien respondio por ello y cuando.
COLUMNAS = [
    ('codigo', 'Codigo del documento'),
    ('nombre', 'Nombre del fichero'),
    ('ruta', 'Ubicacion en el ECD'),
    ('estado', 'Estado ISO 19650'),
    ('idoneidad', 'Codigo de idoneidad'),
    ('revision', 'Revision'),
    ('version', 'Version'),
    ('emitida_en', 'Fecha de emision'),
    ('emitida_por', 'Emitida por'),
    ('subida_en', 'Fecha de subida'),
    ('subida_por', 'Subida por'),
    ('nomenclatura', 'Cumple nomenclatura'),
    ('tamano_mb', 'Tamano (MB)'),
]

# El indice cubre lo que ha salido del trabajo en curso. Un documento en WIP no
# se lista: no se ha comprometido con nadie y meterlo daria una foto falsa de lo
# entregado. Se puede pedir 'TODOS' explicitamente para la foto interna.
ESTADOS_ENTREGADOS = ('SHARED', 'PUBLISHED', 'ARCHIVED')


def filas_del_indice(cursor, model_urn, estados=None):
    """Las filas del indice. Solo lee.

    `estados`: tupla de estados a incluir, o None para los entregados. Pasar
    'TODOS' incluye tambien el trabajo en curso.
    """
    if estados == 'TODOS':
        filtro_estado, params_estado = '', []
    else:
        filtro_estado = 'AND fn.status = ANY(%s)'
        params_estado = [list(estados or ESTADOS_ENTREGADOS)]

    cursor.execute("""
        WITH RECURSIVE rutas AS (
            SELECT id, parent_id, name, CAST(name AS TEXT) AS ruta
              FROM file_nodes
             WHERE parent_id IS NULL AND model_urn = %s
            UNION ALL
            SELECT fn.id, fn.parent_id, fn.name, r.ruta || '/' || fn.name
              FROM file_nodes fn
              INNER JOIN rutas r ON fn.parent_id = r.id
             WHERE fn.model_urn = %s
        )
        SELECT fn.name, r.ruta, fn.status, fn.codigo_idoneidad, fn.codigo_revision,
               fn.version_number, fv.emitida_en, fv.emitida_por,
               fv.created_at, fv.created_by, fn.nomenclatura_ok, fn.size_bytes
          FROM file_nodes fn
          JOIN rutas r ON fn.id = r.id
          -- La version VIGENTE, no la ultima subida: son distintas si alguien
          -- subio algo nuevo despues de la que se aprobo.
          LEFT JOIN file_versions fv ON fv.id = fn.current_version_id
         WHERE fn.model_urn = %s
           AND fn.node_type = 'FILE'
           AND COALESCE(fn.is_deleted, FALSE) = FALSE
           """ + filtro_estado + """
         ORDER BY r.ruta ASC, fn.name ASC
    """, [model_urn, model_urn, model_urn] + params_estado)

    filas = []
    for (nombre, ruta, estado, idoneidad, revision, version,
         emitida_en, emitida_por, subida_en, subida_por, nom_ok, tam) in cursor.fetchall():
        filas.append({
            # El codigo es el nombre sin extension: en esta convencion el nombre
            # del fichero ES el identificador del contenedor de informacion.
            'codigo': (nombre or '').rsplit('.', 1)[0],
            'nombre': nombre,
            # La ruta que se guarda empieza por la raiz interna del arbol, que al
            # usuario no le dice nada. Se quita para que la columna sea legible.
            'ruta': '/'.join((ruta or '').split('/')[1:-1]),
            'estado': estado,
            'idoneidad': idoneidad,
            'revision': revision,
            'version': version,
            'emitida_en': emitida_en.date().isoformat() if emitida_en else None,
            'emitida_por': emitida_por,
            'subida_en': subida_en.date().isoformat() if subida_en else None,
            'subida_por': subida_por,
            'nomenclatura': None if nom_ok is None else ('si' if nom_ok else 'no'),
            'tamano_mb': round((tam or 0) / 1048576.0, 2) if tam else None,
        })
    return filas


def a_excel(filas, model_urn, generado_por=None, estados=None):
    """La hoja de calculo. Devuelve los bytes del .xlsx.

    Sin formulas ni macros: tiene que abrirse en cualquier sitio, hoy y dentro de
    diez anos, y una entidad no acepta un fichero que ejecuta cosas.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Indice del expediente'

    cabecera = Font(bold=True, color='FFFFFF')
    fondo = PatternFill('solid', fgColor='1C4E80')

    # Portada minima: sin esto, una hoja suelta no dice de que obra es ni a que
    # fecha corresponde, y en una auditoria eso la invalida.
    ws.append(['INDICE DEL EXPEDIENTE'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append(['Obra', model_urn])
    ws.append(['Generado', datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')])
    ws.append(['Generado por', generado_por or '-'])
    ws.append(['Alcance', 'todos los estados' if estados == 'TODOS'
               else 'compartido, publicado y archivado (no incluye trabajo en curso)'])
    ws.append(['Documentos', len(filas)])
    ws.append([])

    fila_cab = ws.max_row + 1
    ws.append([etiqueta for _clave, etiqueta in COLUMNAS])
    for c in range(1, len(COLUMNAS) + 1):
        celda = ws.cell(row=fila_cab, column=c)
        celda.font = cabecera
        celda.fill = fondo
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for f in filas:
        ws.append([f.get(clave) for clave, _e in COLUMNAS])

    anchos = {'codigo': 34, 'nombre': 38, 'ruta': 40, 'estado': 14, 'idoneidad': 12,
              'revision': 10, 'version': 8, 'emitida_en': 13, 'emitida_por': 20,
              'subida_en': 13, 'subida_por': 20, 'nomenclatura': 12, 'tamano_mb': 11}
    for i, (clave, _e) in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[ws.cell(row=fila_cab, column=i).column_letter].width = anchos.get(clave, 16)
    ws.freeze_panes = ws.cell(row=fila_cab + 1, column=1)

    salida = io.BytesIO()
    wb.save(salida)
    return salida.getvalue()


def resumen(filas):
    """Cuatro numeros para la pantalla, sin recorrer la tabla a ojo."""
    from collections import Counter
    return {
        'documentos': len(filas),
        'por_estado': dict(Counter(f['estado'] or 'sin estado' for f in filas)),
        'por_idoneidad': dict(Counter(f['idoneidad'] or 'sin codigo' for f in filas)),
        # Lo que le falta al expediente para estar completo. Es el numero que
        # importa antes de una entrega, y el que nadie mira hasta que es tarde.
        'sin_idoneidad': sum(1 for f in filas if not f['idoneidad']),
        'sin_revision': sum(1 for f in filas if not f['revision']),
        'sin_emisor': sum(1 for f in filas if not f['emitida_por']),
    }
