"""Pure parsing and validation helpers for the 4D LOB data pipeline."""
from __future__ import annotations

import hashlib
import io
import re
import xml.etree.ElementTree as ET


MAX_EXCEL_BYTES = 80 * 1024 * 1024
MAX_XML_BYTES = 500 * 1024 * 1024


def number(value):
    try:
        if value is None or value == '':
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def text(value):
    value = str(value).strip() if value is not None else ''
    return value or None


def sha256_bytes(data):
    return hashlib.sha256(data).hexdigest()


def sha256_stream(stream, chunk_size=1024 * 1024):
    digest = hashlib.sha256()
    size = 0
    stream.seek(0)
    while True:
        chunk = stream.read(chunk_size)
        if not chunk:
            break
        size += len(chunk)
        digest.update(chunk)
    stream.seek(0)
    return digest.hexdigest(), size


def validate_source(file_storage, source_type):
    if not file_storage:
        return
    name = (file_storage.filename or '').lower()
    limits = {'duraciones': MAX_EXCEL_BYTES, 'metrados': MAX_EXCEL_BYTES, 'cronograma': MAX_XML_BYTES}
    allowed = {
        'duraciones': ('.xlsm', '.xlsx'),
        'metrados': ('.xlsx', '.xlsm'),
        'cronograma': ('.xml',),
    }
    if not name.endswith(allowed[source_type]):
        raise ValueError(f"Formato invalido para {source_type}: {file_storage.filename}")
    _, size = sha256_stream(file_storage.stream)
    if size <= 0:
        raise ValueError(f"El archivo {file_storage.filename} esta vacio.")
    if size > limits[source_type]:
        raise ValueError(f"El archivo {file_storage.filename} supera el limite permitido.")


def parse_duraciones(file_bytes):
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if 'EDT' not in workbook.sheetnames:
        workbook.close()
        raise ValueError("El archivo de duraciones no tiene hoja 'EDT'.")

    rows = {}
    sheet = workbook['EDT']
    for row in sheet.iter_rows(min_row=5, max_col=87, values_only=True):
        codigo = text(row[23])
        if not codigo:
            continue
        rows[codigo] = {
            'codigo': codigo,
            'descripcion': text(row[24]),
            'unidad': text(row[25]),
            'metrado': number(row[26]),
            'pu': number(row[27]),
            'rendimiento': number(row[36]),
            'duracion': number(row[86]),
            'activity_id': text(row[22]),
            'frente_label': text(row[12]),
            'orden': int(number(row[15]) or 0),
            'tipo': 'partida',
        }
    workbook.close()
    return list(rows.values())


def parse_metrados(file_bytes):
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    control = {}
    titles = {}
    progress = {}
    fronts = []

    if 'CONTROL_OBRA' in workbook.sheetnames:
        for row in workbook['CONTROL_OBRA'].iter_rows(min_row=2, max_col=9, values_only=True):
            codigo = text(row[0])
            tipo = text(row[1])
            if not codigo:
                continue
            if tipo and tipo.upper() == 'TITULO':
                titles[codigo] = text(row[2])
                continue
            replanteado = number(row[5])
            contractual = number(row[4])
            control[codigo] = {
                'descripcion': text(row[2]),
                'unidad': text(row[3]),
                'metrado': replanteado if replanteado is not None else contractual,
                'pu': number(row[6]),
            }

    if 'VALORIZACIONES' in workbook.sheetnames:
        for row in workbook['VALORIZACIONES'].iter_rows(min_row=5, max_col=12, values_only=True):
            codigo = text(row[0])
            if not codigo or not text(row[2]):
                continue
            values = {}
            for offset in range(7):
                value = number(row[4 + offset])
                if value is not None and value != 0:
                    values[offset + 1] = value
            if values:
                progress[codigo] = values

    if 'MAPEO_FRENTES' in workbook.sheetnames:
        for row in workbook['MAPEO_FRENTES'].iter_rows(min_row=2, max_col=2, values_only=True):
            front = text(row[0])
            code = text(row[1])
            if front and code:
                fronts.append({'frente': front, 'cod_base': code})

    workbook.close()
    return {'control': control, 'avance': progress, 'frentes': fronts, 'titulos': titles}


def _local_name(tag):
    return tag.rsplit('}', 1)[-1]


def _p6_fields(element):
    """Return direct P6 activity fields independent of the P6 XML version."""
    return {
        _local_name(child.tag): (child.text or '').strip() or None
        for child in element
    }


def _date(value):
    return value[:10] if value else None


def parse_p6_package(file_stream):
    """Stream a P6 XML export without loading the full document in memory.

    Los exports de P6 traen el proyecto VIGENTE y, aparte, un <BaselineProject>
    (linea base) con las MISMAS actividades pero fechas de baseline. Si no se
    ignora, sus actividades pisan al cronograma real (misma Id) y todo queda
    'Not Started' con fechas de baseline. Aqui se saltan las de baseline.
    Se usa StartDate/FinishDate (cronograma vigente = refleja avances reales),
    cayendo a PlannedStartDate/PlannedFinishDate solo si faltan.
    """
    activities = {}
    object_to_activity = {}
    raw_relations = []
    file_stream.seek(0)
    baseline_depth = 0
    baseline_tags = ('BaselineProject', 'ProjectBaseline')
    for event, element in ET.iterparse(file_stream, events=('start', 'end')):
        name = _local_name(element.tag)
        if event == 'start':
            if name in baseline_tags:
                baseline_depth += 1
            continue
        # event == 'end'
        if name in baseline_tags:
            baseline_depth = max(0, baseline_depth - 1)
            element.clear()
            continue
        if name == 'Relationship' and baseline_depth == 0:
            fields = _p6_fields(element)
            predecessor = fields.get('PredecessorActivityObjectId')
            successor = fields.get('SuccessorActivityObjectId')
            if predecessor and successor:
                raw_relations.append({
                    'predecessor_object_id': predecessor,
                    'successor_object_id': successor,
                    'relation_type': fields.get('Type') or 'Finish to Start',
                    'lag_hours': number(fields.get('Lag')) or 0,
                })
            element.clear()
            continue
        if name != 'Activity':
            continue
        if baseline_depth > 0:  # actividad dentro del BaselineProject → ignorar
            element.clear()
            continue
        fields = _p6_fields(element)
        activity_id = fields.get('Id')
        if activity_id:
            object_id = fields.get('ObjectId')
            if object_id:
                object_to_activity[object_id] = activity_id
            start = _date(fields.get('StartDate')) or _date(fields.get('PlannedStartDate'))
            finish = _date(fields.get('FinishDate')) or _date(fields.get('PlannedFinishDate'))
            pct = number(fields.get('PercentComplete')) or 0
            if 0 < pct <= 1:  # P6 exporta 0..1 (fraccion); normalizar a 0..100
                pct *= 100
            activities[activity_id] = {
                'nombre': fields.get('Name'),
                'start': start,
                'finish': finish,
                'actual_start': _date(fields.get('ActualStartDate')),
                'actual_finish': _date(fields.get('ActualFinishDate')),
                'percent': max(0, min(100, pct)),
                'status': fields.get('Status'),
            }
        element.clear()
    file_stream.seek(0)
    type_map = {
        'finish to start': 'FS', 'start to start': 'SS',
        'finish to finish': 'FF', 'start to finish': 'SF',
        'task dependent': 'FS', 'resource dependent': 'FS',
    }
    relations = []
    for relation in raw_relations:
        predecessor = object_to_activity.get(relation['predecessor_object_id'])
        successor = object_to_activity.get(relation['successor_object_id'])
        if not predecessor or not successor or predecessor == successor:
            continue
        relations.append({
            'predecessor_id': predecessor,
            'successor_id': successor,
            'relation_type': type_map.get(str(relation['relation_type']).lower(), 'FS'),
            'lag_hours': relation['lag_hours'],
            'source': 'p6',
        })
    return {'activities': activities, 'relations': relations}


def parse_p6_xml(file_stream):
    """Compatibility API returning only the current P6 activities."""
    return parse_p6_package(file_stream)['activities']


def merge_cost_items(durations, quantities):
    by_code = {row['codigo']: dict(row) for row in durations}
    for codigo, values in (quantities.get('control') or {}).items():
        item = by_code.setdefault(codigo, {
            'codigo': codigo,
            'descripcion': None,
            'unidad': None,
            'metrado': None,
            'pu': None,
            'rendimiento': None,
            'duracion': None,
            'activity_id': None,
            'frente_label': None,
            'orden': 0,
            'tipo': 'partida',
        })
        item['descripcion'] = item.get('descripcion') or values.get('descripcion')
        item['unidad'] = item.get('unidad') or values.get('unidad')
        if values.get('metrado') is not None:
            item['metrado'] = values['metrado']
        if values.get('pu') is not None:
            item['pu'] = values['pu']

    for codigo, description in (quantities.get('titulos') or {}).items():
        if codigo not in by_code:
            by_code[codigo] = {
                'codigo': codigo,
                'descripcion': description,
                'unidad': None,
                'metrado': None,
                'pu': None,
                'rendimiento': None,
                'duracion': None,
                'activity_id': None,
                'frente_label': None,
                'orden': 0,
                'tipo': 'titulo',
            }

    for codigo, item in by_code.items():
        parts = codigo.split('.')
        item['parent_codigo'] = '.'.join(parts[:-1]) if len(parts) > 1 else None
    return by_code


def quality_report(items, activities, progress, links=0, linked_elements=0):
    actual_items = [item for item in items.values() if item.get('tipo') == 'partida']
    item_codes = {item['codigo'] for item in actual_items}
    with_activity = [item for item in actual_items if item.get('activity_id')]
    matched_activity = [item for item in with_activity if item['activity_id'] in activities]
    progress_codes = item_codes.intersection(progress.keys())

    def pct(value, total):
        return round((value / total) * 100, 2) if total else 0

    warnings = []
    if actual_items and pct(len(matched_activity), len(actual_items)) < 80:
        warnings.append('Menos del 80% de partidas tiene una actividad P6 valida.')
    if actual_items and pct(len(progress_codes), len(actual_items)) < 20:
        warnings.append('Menos del 20% de partidas tiene avance valorizado.')

    return {
        'partidas': len(actual_items),
        'titulos': max(0, len(items) - len(actual_items)),
        'actividades_p6': len(activities),
        'partidas_con_activity_id': len(with_activity),
        'partidas_con_p6_valido': len(matched_activity),
        'partidas_con_avance': len(progress_codes),
        'cobertura_p6_pct': pct(len(matched_activity), len(actual_items)),
        'cobertura_avance_pct': pct(len(progress_codes), len(actual_items)),
        'vinculos_bim': links,
        'elementos_bim_vinculados': linked_elements,
        'warnings': warnings,
    }


def normalize_scope(value):
    value = str(value or '').strip()
    if not value or value.lower() == 'global':
        raise ValueError('Se requiere un proyecto/frente explicito para 4D LOB.')
    if len(value) > 255 or not re.match(r'^[A-Za-z0-9_.:/-]+$', value):
        raise ValueError('El identificador de proyecto/frente no es valido.')
    return value
